import pandas as pd
import os
import glob
import json
import urllib.request
import urllib.error
from datetime import datetime
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, Reference

# Sökvägar baserade på docs/qa/requirements-model-workflow.md
BASE_DIR = os.path.join(os.path.dirname(__file__), '..')
INPUT_DIR = os.path.join(BASE_DIR, 'docs', 'qa', 'requirements-model')
OUTPUT_DIR = os.path.join(BASE_DIR, 'Sammanställning')

NOTEBOOKLM_MAX_CHARS = 900

def read_requirements_csv(path):
    """
    requirements-model exporterar semikolonseparerade CSV-filer.
    Läser med fast delimiter för att undvika pandas-standard (komma) tokenizing-fel.
    """
    return pd.read_csv(path, sep=';', encoding='utf-8')

def normalize_text(value):
    """Normalize values for text export and remove placeholder NaN strings."""
    if pd.isna(value):
        return ""
    text = " ".join(str(value).split()).strip()
    if text.lower() == "nan":
        return ""
    return text

def truncate_text(text, max_chars=NOTEBOOKLM_MAX_CHARS):
    """Truncate long text for better NotebookLM readability."""
    if len(text) <= max_chars:
        return text, False
    return text[: max_chars - 3].rstrip() + "...", True

def load_env_simple():
    """Laddar .env eller .env.local för att hitta API-nycklar utan extra libraries."""
    for fname in ['.env.local', '.env']:
        path = os.path.join(BASE_DIR, fname)
        if os.path.exists(path):
            print(f"Laddar miljövariabler från {fname}...")
            with open(path, 'r', encoding='utf-8') as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith('#') and '=' in line:
                        key, value = line.split('=', 1)
                        if key not in os.environ:
                            os.environ[key] = value.strip().strip('"\'')
            break

def call_gemini_summary(text_content):
    """Skapar koppling mot Google Gemini för att analysera datan."""
    api_key = os.environ.get("GEMINI_API_KEY")
    if not api_key:
        print("OBS: Ingen GEMINI_API_KEY hittades. Hoppar över AI-koppling.")
        return None

    # Använder Gemini 1.5 Flash för snabb analys av stora textmängder
    url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash-latest:generateContent?key={api_key}"
    headers = {"Content-Type": "application/json"}
    prompt = "Du är en dataanalytiker. Analysera denna sammanställning av myndighetskrav. Identifiera 3 tydliga trender och 1 avvikelse:\n\n" + text_content[:40000]
    
    data = {"contents": [{"parts": [{"text": prompt}]}]}
    
    try:
        req = urllib.request.Request(url, data=json.dumps(data).encode('utf-8'), headers=headers)
        with urllib.request.urlopen(req) as response:
            result = json.loads(response.read().decode('utf-8'))
            return result.get('candidates', [{}])[0].get('content', {}).get('parts', [{}])[0].get('text', '')
    except urllib.error.HTTPError as e:
        print(f"Kopplingsfel mot Google Gemini (HTTP {e.code}): {e.reason}")
        try:
            print(e.read().decode('utf-8'))
        except:
            pass
        return None
    except Exception as e:
        print(f"Oväntat fel vid AI-koppling: {e}")
        return None

def generate_excel():
    print("Läser in CSV-filer från requirements-model...")
    load_env_simple()
    
    # Skapa mapp och rensa gamla filer
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)
        print(f"Skapade mapp: {OUTPUT_DIR}")
        
    print(f"Rensar gamla filer i {OUTPUT_DIR}...")
    for f in glob.glob(os.path.join(OUTPUT_DIR, "Examensmatris_Sammanställning_*.xlsx")):
        try: os.remove(f)
        except: pass
    for f in glob.glob(os.path.join(OUTPUT_DIR, "Examensunderlag_NotebookLM_*.txt")):
        try: os.remove(f)
        except: pass

    output_file = os.path.join(OUTPUT_DIR, f'Examensmatris_Sammanställning_{datetime.now().strftime("%Y-%m-%d_%H%M")}.xlsx')

    try:
        # Läs in Cases (Metadata om ärendet)
        cases_df = read_requirements_csv(os.path.join(INPUT_DIR, 'requirement_cases.csv'))
        
        # Läs in Rows (Specifika kravrader)
        rows_df = read_requirements_csv(os.path.join(INPUT_DIR, 'requirement_rows.csv'))
        
        print(f"DEBUG: Totalt antal ärenden i exportfilen (cases): {len(cases_df)}")
        if 'Kommun' in cases_df.columns:
            print(f"DEBUG: Totalt antal unika kommuner i exportfilen: {cases_df['Kommun'].nunique()}")
        print(f"DEBUG: Totalt antal extraherade kravrader (rows): {len(rows_df)}")

        # Slå ihop tabellerna på CaseId för att få en komplett "Master Matris"
        # Detta gör att varje kravrad nu har information om Kommun, Årtal, Myndighet etc.
        master_df = pd.merge(rows_df, cases_df, on='CaseId', how='left')

        # --- DATATVÄTT (Power Query-liknande steg) ---
        print("Utför datatvätt (rensning av whitespace, datumformat, 'nan'-värden)...")
        
        # 1. Rensa textfält (trimma och ta bort dubbla mellanslag)
        for col in master_df.select_dtypes(include=['object']).columns:
            master_df[col] = master_df[col].apply(lambda x: " ".join(str(x).split()) if pd.notna(x) else "")
            master_df[col] = master_df[col].replace({'nan': '', 'NaN': '', 'None': ''})

        # 1b. Specifik tvätt av Kommun-kolumnen (ta bort sidhuvuden/datum)
        if 'Kommun' in master_df.columns:
            def clean_kommun_field(val):
                if not val: return ""
                # Ta bort siffror (datum, sidnummer, orgnr)
                val = ''.join([c for c in val if not c.isdigit()])
                # Ta bort vanliga skräpord från sidhuvuden
                garbage = ["Sida", "Datum", "Dnr", "Diarienummer", "Beslut", "Anmälan", "Protokoll", "Justering", "Avsändare", "Mottagare", "Sammanträdesdatum"]
                for g in garbage:
                    val = val.replace(g, "").replace(g.upper(), "").replace(g.lower(), "")
                # Ta bort specialtecken
                for char in ['|', '/', ':', ';', '(', ')', '[', ']', '{', '}']:
                    val = val.replace(char, "")
                return " ".join(val.split())

            master_df['Kommun'] = master_df['Kommun'].apply(clean_kommun_field)

        # 2. Formatera datum snyggt (ISO 8601)
        if 'Dokumentdatum' in master_df.columns:
            master_df['Dokumentdatum'] = pd.to_datetime(master_df['Dokumentdatum'], errors='coerce').dt.strftime('%Y-%m-%d')

        # Säkerställ att kolumn för manuell verifiering finns
        if 'VerifieradAv' not in master_df.columns:
            master_df['VerifieradAv'] = ''
        master_df['VerifieradAv'] = master_df['VerifieradAv'].fillna('')

        # Skapa aktiva PDF-länkar om KallaFil finns
        if 'KallaFil' in master_df.columns:
            def create_hyperlink(path):
                if pd.isna(path) or str(path).strip() == "":
                    return ""
                # Skapa absolut sökväg för att länken ska fungera oavsett var Excel-filen sparas
                full_path = path if os.path.isabs(str(path)) else os.path.abspath(os.path.join(BASE_DIR, str(path)))
                # Excel HYPERLINK-formel
                return f'=HYPERLINK("{full_path}", "Öppna PDF")'
            
            master_df['PDF_Länk'] = master_df['KallaFil'].apply(create_hyperlink)
        
        print(f"Skapar Excel-fil: {output_file}")
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # 0. Sammanställning (Dashboard)
            summary_rows = [
                {'Nyckeltal': 'Rapportdatum', 'Värde': datetime.now().strftime("%Y-%m-%d %H:%M")},
                {'Nyckeltal': 'Totalt antal kravrader', 'Värde': len(master_df)},
            ]
            if 'Kommun' in master_df.columns:
                summary_rows.append({'Nyckeltal': 'Antal unika kommuner (med krav)', 'Värde': master_df['Kommun'].nunique()})
            
            if 'Kommun' in cases_df.columns:
                summary_rows.append({'Nyckeltal': 'Antal unika kommuner (totalt i underlag)', 'Värde': cases_df['Kommun'].nunique()})
            
            if 'VerifieradAv' in master_df.columns:
                 verified_count = master_df[master_df['VerifieradAv'] != ''].shape[0]
                 summary_rows.append({'Nyckeltal': 'Antal manuellt verifierade', 'Värde': f"{verified_count} ({int(verified_count/len(master_df)*100)}%)"})

            pd.DataFrame(summary_rows).to_excel(writer, sheet_name='Sammanställning', index=False)

            # 1. Master Data (Källan för dina Pivottabeller)
            master_df.to_excel(writer, sheet_name='Master_Data', index=False)
            
            # 2. Table A: Ärenden per myndighet (Exempel på sammanställning)
            if 'Myndighet' in master_df.columns:
                table_a = master_df.groupby('Myndighet').size().reset_index(name='Antal Krav')
                table_a.to_excel(writer, sheet_name='Table_A_Myndighet', index=False)

            # 3. Table B: Kravfrekvens per kategori
            if 'Kravkategori' in master_df.columns:
                table_b = master_df['Kravkategori'].value_counts().reset_index()
                table_b.columns = ['Kategori', 'Frekvens']
                table_b.to_excel(writer, sheet_name='Table_B_Kategorier', index=False)
                
                # --- CHART (Graf) ---
                workbook = writer.book
                worksheet = writer.sheets['Table_B_Kategorier']
                
                chart = BarChart()
                chart.type = "col"
                chart.style = 10
                chart.title = "Fördelning av Kravkategorier"
                chart.y_axis.title = 'Antal'
                chart.x_axis.title = 'Kategori'

                data = Reference(worksheet, min_col=2, min_row=1, max_row=len(table_b)+1, max_col=2)
                cats = Reference(worksheet, min_col=1, min_row=2, max_row=len(table_b)+1)
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                worksheet.add_chart(chart, "D2")

            # --- FORMATERING AV SAMMANSTÄLLNING ---
            ws_summary = writer.sheets['Sammanställning']
            ws_summary.column_dimensions['A'].width = 35
            ws_summary.column_dimensions['B'].width = 25
            for cell in ws_summary["1:1"]:
                cell.font = Font(bold=True, size=12)

            # 4. Table C: Kommunskillnader (Ytkonstruktion vs Lakvatten)
            if 'Kravkategori' in master_df.columns and 'Kommun' in master_df.columns:
                # Filtrera fram relevanta kategorier
                relevant_cats = ['Ytkonstruktion', 'DagvattenLakvatten']
                table_c = master_df[master_df['Kravkategori'].isin(relevant_cats)]
                # Pivotera för att se fördelning per kommun
                table_c_pivot = pd.crosstab(table_c['Kommun'], table_c['Kravkategori'])
                table_c_pivot.to_excel(writer, sheet_name='Table_C_Kommun_Jmf')

            # 5. Table D: Avfallskoder (EWC)
            # Antar att det finns en kolumn för EWC eller Avfallstyp, annars hoppar vi över
            if 'Avfallskod' in master_df.columns:
                table_d = master_df['Avfallskod'].value_counts().reset_index(name='Antal')
                table_d.to_excel(writer, sheet_name='Table_D_EWC', index=False)

            # 6. Table E: Län/Region (Om data finns)
            if 'Lan' in master_df.columns:
                table_e = master_df['Lan'].value_counts().reset_index(name='Antal')
                table_e.to_excel(writer, sheet_name='Table_E_Lan', index=False)
                
                # Skapa graf för Län om det finns data
                ws_e = writer.sheets['Table_E_Lan']
                chart_e = BarChart()
                chart_e.type = "col"
                chart_e.style = 10
                chart_e.title = "Fördelning per Län"
                chart_e.y_axis.title = 'Antal'
                chart_e.x_axis.title = 'Län'
                
                data_e = Reference(ws_e, min_col=2, min_row=1, max_row=len(table_e)+1, max_col=2)
                cats_e = Reference(ws_e, min_col=1, min_row=2, max_row=len(table_e)+1)
                chart_e.add_data(data_e, titles_from_data=True)
                chart_e.set_categories(cats_e)
                ws_e.add_chart(chart_e, "D2")

            # 7. Table F: Detaljerad specifikation (Drill-down)
            # Gruppera på Kategori -> Subkategori -> Avfallsslag för att se detaljerna
            drill_cols = []
            if 'Kravkategori' in master_df.columns: drill_cols.append('Kravkategori')
            if 'Kravsubkategori' in master_df.columns: drill_cols.append('Kravsubkategori')
            if 'Avfallsslag' in master_df.columns: drill_cols.append('Avfallsslag')
            
            if len(drill_cols) > 1:
                table_f = master_df.groupby(drill_cols).size().reset_index(name='Antal Krav')
                table_f = table_f.sort_values(by=['Antal Krav'], ascending=False)
                table_f.to_excel(writer, sheet_name='Table_F_Detaljanalys', index=False)

            # --- FORMATERING AV MASTER DATA ---
            workbook = writer.book
            worksheet = writer.sheets['Master_Data']
            
            # Hämta dimensioner
            max_row = len(master_df) + 1
            max_col = len(master_df.columns)
            max_col_letter = get_column_letter(max_col)
            full_range = f"A1:{max_col_letter}{max_row}"

            # 1. SKAPA EN RIKTIG EXCEL-TABELL (Gör det enkelt att filtrera/sortera)
            # Detta gör att du kan klicka "Infoga Pivottabell" i Excel och den väljer allt automatiskt.
            tab = Table(displayName="KravData", ref=full_range)
            style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            tab.tableStyleInfo = style
            worksheet.add_table(tab)

            # Hitta vilken kolumn (bokstav) som är 'VerifieradAv'
            if 'VerifieradAv' in master_df.columns:
                col_idx = master_df.columns.get_loc('VerifieradAv') + 1
                col_letter = get_column_letter(col_idx)
                
                # 2. VILLKORLIG FORMATERING (Grönmarkera verifierade rader)
                green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                range_string = f"A2:{max_col_letter}{max_row}"
                rule = FormulaRule(formula=[f'${col_letter}2<>""'], stopIfTrue=True, fill=green_fill)
                worksheet.conditional_formatting.add(range_string, rule)

                # 3. DATA VALIDATION (Dropdown-lista för verifiering)
                # Skapa en dropdown i 'VerifieradAv'-kolumnen
                dv = DataValidation(type="list", formula1='"OK,Jimmy,Admin,Ej Relevant"', allow_blank=True)
                dv.error = 'Välj ett värde från listan'
                dv.errorTitle = 'Ogiltigt val'
                dv.prompt = 'Välj vem som verifierat raden'
                dv.promptTitle = 'Verifiering'
                
                # Applicera på hela kolumnen (exklusive rubrik)
                dv.add(f"{col_letter}2:{col_letter}{max_row}")
                worksheet.add_data_validation(dv)

            # Justera kolumnbredder lite (enkel heuristik)
            for i, column in enumerate(master_df.columns):
                # Hantera NaN/icke-strängvärden robust vid längdberäkning
                value_max_len = master_df[column].map(lambda value: len(str(value)) if pd.notna(value) else 0).max()
                column_len = max(value_max_len, len(str(column))) + 2
                # Sätt en maxbredd så inte beskrivningstexter tar över hela skärmen
                if column_len > 50: column_len = 50
                worksheet.column_dimensions[get_column_letter(i+1)].width = column_len

        print("Klart! Excel-filen är skapad.")
        print("Datan är nu formaterad som en Excel-tabell ('KravData').")
        print("För att skapa en Pivottabell: Klicka var som helst i tabellen -> Infoga -> Pivottabell.")

        # --- GENERERA AI-PROMPT UNDERLAG ---
        print("\n" + "="*50)
        print("  DATA-SAMMANFATTNING FÖR AI-PROMPT  ")
        print("="*50)
        print(f"Totalt antal analyserade krav: {len(master_df)}")
        if 'Kravkategori' in master_df.columns:
            print("\nTopp 5 Kravkategorier (använd i Resultat-kapitlet):")
            cats = master_df['Kravkategori'].value_counts(normalize=True) * 100
            for cat, pct in cats.head(5).items():
                print(f"- {cat}: {pct:.1f}%")
        
        if 'Kravniva' in master_df.columns:
            print("\nFördelning SKA vs BÖR krav:")
            levels = master_df['Kravniva'].value_counts(normalize=True) * 100
            for lvl, pct in levels.items():
                print(f"- {lvl}: {pct:.1f}%")

        if 'Lan' in master_df.columns:
            print("\nTopp 5 Län:")
            lan_counts = master_df['Lan'].value_counts(normalize=True) * 100
            for lan, pct in lan_counts.head(5).items():
                print(f"- {lan}: {pct:.1f}%")
        
        if 'Kravsubkategori' in master_df.columns:
            print("\nTopp 5 Subkategorier (Detaljnivå):")
            subcats = master_df['Kravsubkategori'].value_counts(normalize=True) * 100
            for sub, pct in subcats.head(5).items():
                print(f"- {sub}: {pct:.1f}%")
        print("="*50 + "\n")

        # --- GENERERA TEXTFIL FÖR NOTEBOOKLM ---
        txt_output = os.path.join(OUTPUT_DIR, f'Examensunderlag_NotebookLM_{datetime.now().strftime("%Y-%m-%d_%H%M")}.txt')
        print(f"Skapar textunderlag optimerat för NotebookLM: {txt_output}")
        
        exported_rows = 0
        truncated_rows = 0
        with open(txt_output, 'w', encoding='utf-8') as f:
            f.write("# SAMMANSTÄLLNING AV MYNDIGHETSKRAV (KÄLLDATA)\n")
            f.write(f"Genererad: {datetime.now().strftime('%Y-%m-%d')}\n")
            f.write("Detta dokument innehåller alla extraherade kravtexter sorterade per kategori.\n")
            f.write("Använd detta för att fråga NotebookLM om kvalitativa mönster.\n\n")

            # --- NYTT: Skriv in statistiken i textfilen så NotebookLM slipper räkna själv ---
            f.write("## 1. STATISTISK ÖVERSIKT (FACIT)\n")
            f.write("Använd dessa siffror när du refererar till mängder och fördelning.\n\n")

            if 'Kravkategori' in master_df.columns:
                f.write("### Tabell: Kravfördelning per Kategori\n")
                f.write("| Kategori | Antal |\n|---|---|\n")
                for cat, count in master_df['Kravkategori'].value_counts().items():
                    f.write(f"| {cat} | {count} |\n")
                f.write("\n")

            if 'Lan' in master_df.columns:
                f.write("### Tabell: Geografisk fördelning (Län)\n")
                f.write("| Län | Antal |\n|---|---|\n")
                for lan, count in master_df['Lan'].value_counts().items():
                    f.write(f"| {lan} | {count} |\n")
                f.write("\n")

            drill_cols = []
            if 'Kravkategori' in master_df.columns: drill_cols.append('Kravkategori')
            if 'Kravsubkategori' in master_df.columns: drill_cols.append('Kravsubkategori')
            
            if len(drill_cols) > 1:
                f.write("### Tabell: Detaljerad specifikation (Topp 20)\n")
                f.write("| Kategori > Subkategori | Antal |\n|---|---|\n")
                # Gruppera och rÃ¤kna
                detail_counts = master_df.groupby(drill_cols).size().sort_values(ascending=False).head(20)
                for idx, count in detail_counts.items():
                    label = " > ".join(map(str, idx)) if isinstance(idx, tuple) else str(idx)
                    f.write(f"| {label} | {count} |\n")
                f.write("\n")
            
            if 'Kravkategori' in master_df.columns:
                # Sortera per kategori för tydlighet
                for cat, group in master_df.groupby('Kravkategori'):
                    f.write(f"## KATEGORI: {str(cat).upper()}\n")
                    f.write(f"Antal krav i denna kategori: {len(group)}\n\n")
                    
                    for idx, row in group.iterrows():
                        kommun = normalize_text(row.get('Kommun', 'Okänd')) or 'Okänd'
                        typ = normalize_text(row.get('KravkallaTyp', 'Krav')) or 'Krav'
                        text = normalize_text(row.get('KravtextCitat', ''))

                        if text:
                            text, was_truncated = truncate_text(text)
                            truncated_tag = " [TRUNCATED]" if was_truncated else ""
                            f.write(f"* **{kommun}** ({typ}){truncated_tag}: \"{text}\"\n")
                            exported_rows += 1
                            if was_truncated:
                                truncated_rows += 1
                    f.write("\n" + "-"*40 + "\n\n")

            f.write("## EXPORTINFO\n")
            f.write(f"- Max tecken per kravtext: {NOTEBOOKLM_MAX_CHARS}\n")
            f.write(f"- Exporterade kravrader: {exported_rows}\n")
            f.write(f"- Trunkerade kravrader: {truncated_rows}\n")

        print(f"NotebookLM-export klar. Exporterade rader: {exported_rows}, trunkerade: {truncated_rows}.")

        # --- KOPPLING MOT GOOGLE GEMINI ---
        print("\n" + "-"*40)
        print("Kör koppling mot Google Gemini (AI-analys)...")
        with open(txt_output, 'r', encoding='utf-8') as f:
            full_text = f.read()
        
        ai_summary = call_gemini_summary(full_text)
        if ai_summary:
            ai_file = os.path.join(OUTPUT_DIR, f'AI_Analys_{datetime.now().strftime("%Y-%m-%d_%H%M")}.md')
            with open(ai_file, 'w', encoding='utf-8') as f:
                f.write("# AI-ANALYS AV KRAVBILDEN\n\n")
                f.write(ai_summary)
            print(f"Koppling lyckades! Analys sparad i: {ai_file}")
        else:
            print("Ingen analys kunde genereras.")

    except FileNotFoundError as e:
        print(f"Fel: Kunde inte hitta filen. Har du kört 'npm run requirements:model'?\nDetaljer: {e}")
    except Exception as e:
        print(f"Ett oväntat fel uppstod: {e}")

if __name__ == "__main__":
    generate_excel()
